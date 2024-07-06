from influencer_input_parser import parse_input_data
from urllib.parse import urlparse, parse_qs
import requests
import io
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import json
from tqdm import tqdm
from youtube_video_analysis import RequestChannelInfo, RequestVideoInfo, RequestChannelContentsInfo

# Todo: add error handling, excel cell size
RETURN_ERR = -1


def run_influencer_analysis(args):
    print("[Info] Run Influencer Analysis")
    input_excel, influencer_sheet, video_sheet, start_row, start_col, end_row, max_result, output_excel, dev_keys = parse_input_data(args.input_txt)
    RUN_YC = args.yc
    RUN_YV = args.yv
    
    # read excel
    xlsx = openpyxl.load_workbook(input_excel)
    sheet_num = len(xlsx.sheetnames)
    print("[Info] Open input excel: " + input_excel)

    # run Youtube Channel Analysis
    if ((RUN_YC == True) and (influencer_sheet < sheet_num)):
        cSheet = xlsx.worksheets[influencer_sheet]
        run_InfluencerAnalysis(cSheet, start_row, start_col, end_row, dev_keys, max_result)

    # run Youtube Video Analysis
    if ((RUN_YV == True) and (video_sheet < sheet_num)):
        vSheet = xlsx.worksheets[video_sheet]
        run_VideoAnalysis(vSheet, start_row, start_col, end_row, dev_keys)
    
    # save output excel
    SaveExcel(xlsx, output_excel)
        

def run_VideoAnalysis(sheet, start_row, start_col, end_row, dev_keys):
    print("[Info] Running Youtube Video Analysis")
    max_row = min(sheet.max_row + 1, end_row)
    for row in tqdm(start_row, max_row+ 1):
        vURL = sheet.cell(row, start_col).value
        if vURL == None:
            continue
        vID = get_video_id(vURL)
        if (vID == RETURN_ERR) or (vID == None):
            print("[Warning] " + "fail to get ID from URL : " + vURL)
            continue
        res_json = RequestVideoInfo(vID, dev_keys)
        df_just_video = GetVideoData(vID, res_json, dev_keys)
        if df_just_video == RETURN_ERR:
            continue
        UpdateVideoInfoToExcel(sheet, row, start_col + 1, df_just_video)


def run_InfluencerAnalysis(sheet, start_row, start_col, end_row, dev_keys, max_result):
    print("[Info] Running Youtube Influencer Analysis")
    max_row = min(sheet.max_row + 1, end_row)
    for row in tqdm(start_row, max_row + 1):
        cURL = sheet.cell(row, start_col).value
        if cURL == None:
            print("[Warning] URL = None. row=" + str(row))
            continue
        cID = get_channel_id(cURL)
        if (cID == RETURN_ERR) or (cID == None):
            print("[Warning] " + "fail to get ID from URL : " + cURL)
            continue
            # here!
        channel_info = RequestChannelInfo(cID, dev_keys)
        channel_contents_info = RequestChannelContentsInfo(dev_keys, cID, max_result)
        if channel_contents_info == RETURN_ERR:
            print("[Warning] channel_contents_info = RETURN_ERR")
            continue
        df_just_channel = GetChannelData(cID, channel_info, channel_contents_info, dev_keys)
        if df_just_channel == RETURN_ERR:
            print("[Warning] df_just_channel = RETURN_ERR")
            continue
        UpdateChannelInfoToExcel(sheet, row, start_col + 1, df_just_channel)



#make enum
def make_enum(*sequential, **named):
    enums = dict(zip(sequential, range(len(sequential))), **named)
    return type('Enum', (), enums)

vIndex = make_enum('V_URL', 'V_TITLE', 'VIEW', 'LIKE', 'COMMENTS', 'C_TITLE', 'C_URL', 'CHANNEL_SUBSCRIBER', 'THUMBNAIL')
cIndex = make_enum('URL', 'PROFILE_IMG', 'TITLE', 'SUBSCRIBER', 'POST_VIEW', 'POST_LIKE', 'POST_COMMENT', 'POST_ENGAGE', 'AGE', 'GENDER', 'LOCATION', 'LANGUAGE')


def get_channel_id(youtube_url):
    response = requests.get(youtube_url)
    html_content = response.text
    id_pattern = 'channelId\":\"'
    index = html_content.find(id_pattern)
    channel_id = html_content[index + len(id_pattern):index + len(id_pattern) + 24]
    return channel_id

def get_video_id(youtube_url):
    response = requests.get(youtube_url)
    html_content = response.text
    id_pattern = 'videoId\":\"'
    index = html_content.find(id_pattern)
    video_id = html_content[index + len(id_pattern):index + len(id_pattern) + 11]
    return video_id

def UpdateChannelInfoToExcel(sheet, r, start, data):
    start_c = start - 1
    InsertImage(sheet, data[cIndex.PROFILE_IMG], r, start_c + cIndex.PROFILE_IMG)
    sheet.cell(row=r, column=start_c + cIndex.TITLE).value = '=HYPERLINK("{}", "{}")'.format(data[cIndex.URL], data[cIndex.TITLE])
    sheet.cell(row=r, column=start_c + cIndex.SUBSCRIBER).value = data[cIndex.SUBSCRIBER]

    sheet.cell(row=r, column=start_c + cIndex.POST_VIEW).value = round(float(data[cIndex.POST_VIEW]), 2)
    sheet.cell(row=r, column=start_c + cIndex.POST_LIKE).value = round(float(data[cIndex.POST_LIKE]), 2)
    sheet.cell(row=r, column=start_c + cIndex.POST_COMMENT).value = round(float(data[cIndex.POST_COMMENT]),2 )
    sheet.cell(row=r, column=start_c + cIndex.POST_ENGAGE).value = str(round(float(data[cIndex.POST_ENGAGE]), 2)) + "%"


def UpdateVideoInfoToExcel(sheet, r, start, data):
    start_c = start - 1
    sheet.cell(row=r, column=start_c + vIndex.V_TITLE).value = '=HYPERLINK("{}", "{}")'.format(data[vIndex.V_URL], data[vIndex.V_TITLE])
    sheet.cell(row=r, column=start_c + vIndex.VIEW).value = round(float(data[vIndex.VIEW]), 2)
    sheet.cell(row=r, column=start_c + vIndex.LIKE).value = round(float(data[vIndex.LIKE]), 2)
    sheet.cell(row=r, column=start_c + vIndex.COMMENTS).value = round(float(data[vIndex.COMMENTS]), 2)
    sheet.cell(row=r, column=start_c + vIndex.C_TITLE).value = '=HYPERLINK("{}", "{}")'.format(data[vIndex.C_URL], data[vIndex.C_TITLE])
    sheet.cell(row=r, column=start_c + vIndex.C_URL).value = data[vIndex.C_URL]
    sheet.cell(row=r, column=start_c + vIndex.CHANNEL_SUBSCRIBER).value = round(float(data[vIndex.CHANNEL_SUBSCRIBER]), 2)
    InsertImage(sheet, data[vIndex.THUMBNAIL], r, start_c + vIndex.THUMBNAIL)


def GetChannelData(cID, channel_info, channel_contents_info, dev_keys):
    arr = json.dumps(channel_info)
    jsonObject = json.loads(arr)
    if ((jsonObject.get('error')) or ('items' not in jsonObject)):
        print("[Warning] response error! : " + cID)
        print(jsonObject['error'])
        return RETURN_ERR
    items = jsonObject['items']
    if len(items) <= 0:
        print("[Error] no items for Channel Data: " + cID)
        return RETURN_ERR
    item = jsonObject['items'][0]
    ret = {}
    ret[cIndex.URL] = "https://www.youtube.com/channel/" + cID
    ret[cIndex.PROFILE_IMG] = ""
    ret[cIndex.TITLE] = ""
    ret[cIndex.SUBSCRIBER] = 0
    ret[cIndex.POST_VIEW] = 0
    ret[cIndex.POST_LIKE] = 0
    ret[cIndex.POST_COMMENT] = 0
    ret[cIndex.POST_ENGAGE] = 0
    

    if ('snippet' in item):
        snippet = item['snippet']
        if ('thumbnails' in snippet) and ('high' in snippet['thumbnails']) and ('url' in snippet['thumbnails']['high']):
            ret[cIndex.PROFILE_IMG] = snippet['thumbnails']['high']['url']
        if ('title' in snippet):
            ret[cIndex.TITLE] = snippet['title']

    if ('statistics' in item) and ('subscriberCount' in item['statistics']):
        ret[cIndex.SUBSCRIBER] = item['statistics']['subscriberCount']

    nViewCnt = 0
    nLikeCnt = 0
    nCommentCnt = 0
    nView = 0
    nLike = 0
    nComment = 0
    for content in channel_contents_info.get("items", []):
        if ('id' in content):
            contentId = content['id']
            if ('kind' in contentId) and (contentId["kind"] != "youtube#video"):
                print("[Warning] Type is not video!! check the input: " + cID)
                return RETURN_ERR
            vID = content["id"]["videoId"]
            res_json = RequestVideoInfo(vID, dev_keys)

            video_info = GetVideoData(vID, res_json, dev_keys)
            if (video_info == RETURN_ERR):
                return RETURN_ERR
            view = int(video_info[vIndex.VIEW])
            like = int(video_info[vIndex.LIKE])
            comments = int(video_info[vIndex.COMMENTS])

            if view > 0:
                nViewCnt += view
                nView += 1
            if like > 0:
                nLikeCnt += like
                nLike += 1
            if comments > 0:
                nCommentCnt += comments
                nComment += 1
            break

    if nView > 0:
        ret[cIndex.POST_VIEW] = nViewCnt / nView
    if nLike > 0:
        ret[cIndex.POST_LIKE] = nLikeCnt / nLike
    if nComment > 0:
        ret[cIndex.POST_COMMENT] = nCommentCnt / nComment
    if nViewCnt > 0:
        ret[cIndex.POST_ENGAGE] = ((nLikeCnt + nCommentCnt) / nViewCnt) * 100
    return ret


def GetVideoData(vID, input_json, dev_keys):
    arr = json.dumps(input_json)
    jsonObject = json.loads(arr)
    if ((jsonObject.get('error')) or ('items' not in jsonObject)):
        print("[Warning] response error! : " + vID)
        print(jsonObject['error'])
        return RETURN_ERR
    items = jsonObject['items']
    if len(items) <= 0:
        print("[Warning] no items for Video Data: " + vID)
        return RETURN_ERR
    item = jsonObject['items'][0]
    ret = {}
    ret[vIndex.V_URL] = "https://www.youtube.com/watch?v=" + vID
    ret[vIndex.V_TITLE] = ""
    ret[vIndex.VIEW] = 0
    ret[vIndex.LIKE] = 0
    ret[vIndex.COMMENTS] = 0 
    ret[vIndex.C_TITLE] = ""
    ret[vIndex.C_URL] = ""
    ret[vIndex.CHANNEL_SUBSCRIBER] = 0
    ret[vIndex.THUMBNAIL] = ""    
    
    if ('snippet' in item):
        snippet = item['snippet']
        if ('title' in snippet):
            ret[vIndex.V_TITLE] = snippet['title']
        if ('channelId' in snippet):
            cID = snippet['channelId']
            ret[vIndex.C_URL] = "https://www.youtube.com/channel/" + cID
            channel_info = RequestChannelInfo(cID, dev_keys)
            arr = json.dumps(channel_info)
            jsonObject = json.loads(arr)
            if ('items' in jsonObject):
                item_channel = jsonObject['items'][0]
                if ('snippet' in item_channel) and ('title' in item_channel['snippet']):
                    ret[vIndex.C_TITLE] = item_channel['snippet']['title']
                if ('statistics' in item_channel) and ('subscriberCount' in item_channel['statistics']):
                    ret[vIndex.CHANNEL_SUBSCRIBER] = item_channel['statistics']['subscriberCount']
    if ('statistics' in item):
        statistics = item['statistics']
        if ('viewCount' in statistics):
            ret[vIndex.VIEW] = statistics['viewCount']
        if ('likeCount' in statistics):
            ret[vIndex.LIKE] = statistics['likeCount']
        if ('commentCount' in statistics):
            ret[vIndex.COMMENTS] = statistics['commentCount']
    ret[vIndex.THUMBNAIL] = "https://img.youtube.com/vi/" + vID + "/maxresdefault.jpg"

    return ret


#make output
def InsertImage(sheet, img_url, row, col):
    image_scale = 10
    if (img_url == ""):
        return
    response = requests.get(img_url)
    img_file = io.BytesIO(response.content)
    thumbnailImage = Image(img_file)
    thumbnailImage.width /= image_scale
    thumbnailImage.height /= image_scale
    colChar = get_column_letter(col)
    thumbnailImage.anchor = "%s"%colChar + "%s"%row
    sheet.add_image(thumbnailImage)
    #sheet.column_dimensions[colChar].width = thumbnailImage.width
    sheet.row_dimensions[row].height = thumbnailImage.height

def SaveExcel(xlsx, output_excel):
    xlsx.save(output_excel)
    print("[Info] Done saving excel: " + output_excel)
