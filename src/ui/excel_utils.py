from typing import List

import openpyxl


def parse_channel_urls_from_excel(path: str) -> List[str]:
    """Read channel URLs out of the column headed "URL" on the first sheet of an input workbook."""
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook.worksheets[0]

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    url_col = None
    for index, header in enumerate(header_row):
        if isinstance(header, str) and header.strip().upper() == "URL":
            url_col = index
            break
    if url_col is None:
        return []

    urls = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if url_col < len(row):
            value = row[url_col]
            if isinstance(value, str) and value.strip():
                urls.append(value.strip())
    return urls
